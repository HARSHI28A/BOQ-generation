from flask import Flask, render_template, request
import openpyxl
from flask import send_file


app = Flask(__name__)
list1 = []
list2 = []

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/item_rate.html')
def item_rate():
    return render_template('item_rate.html')

@app.route('/item_rate.html',methods=['POST'])
def getvalue():
    print('1')
    
    items = request.form.getlist('mc')
    items2 = request.form.getlist('bc')
    #list1 = []
    #print(wb.sheetnames)
    #ws['A22'] = 'harshita'
    print(items)
    print(items2)
    wb2 = openpyxl.load_workbook(filename = "templates/Book1.xlsx")
    ws2 = wb2.active
    ws2.column_dimensions.group('B','BC', hidden=True)


    if len(items) != 0:
        for x in items:
            if x=='item_description':
                list1.append('B13')
                list2.append('B')
            elif x == 'item_code':
                list1.append('C13')
                list2.append('C')
            elif x == 'quantity':
                list1.append('D13')
                list2.append('D')
            elif x == 'units':
                list1.append('E13')
                list2.append('E')
            elif x == 'estimated_rate':
                list1.append('F13')
                list2.append('F')
            elif x == 'addition_values':
                list1.append('J13')
                list2.append('J')
            elif x =='addition':
                list1.append('I13')
                list2.append('I')
            elif x == 'currency_conversation':
                list1.append('K13')
                list2.append('K')
            elif x == 'quoted_currency':
                list1.append('L13')
                list2.append('L')
            elif x== 'basic_rate':
                list1.append('M13')
                list2.append('M')
            elif x == 'gst_amount':
                list1.append('R13')
                list2.append('R')
            elif x == 'freight_charges':
                list1.append('P13')
                list2.append('P')
            elif x == 'total_without':
                list1.append('BA13')
                list2.append('BA')
            elif x == 'total_with':
                list1.append('BB13')
                list2.append('BB')

    for kl in list2:
            ws2.column_dimensions[kl].hidden = False

    if len(list1) == len(items2):
        print("Kl")
        print(list2)
        for il , jl in zip(list1, items2):
            ws2[il] = jl
        return render_template('pass2.html')

    wb2.save("templates/Book1.xlsx")   
    print(list1)
    print(list2)
    return render_template('pass.html', items= items)

@app.route('/',methods=['POST'])
def do_stuff():
    data = request.form.getlist('name')
    print(data)
    return ('done')

#@app.route('/Book1.xlsx')
#def downloadFile ():
    #For windows you need to use drive name [ex: F:/Example.pdf]
    #path = "/Book1.xlsx"
    #return send_file(path, as_attachment=True)






if __name__== '__main__':
    app.run(debug=True)
